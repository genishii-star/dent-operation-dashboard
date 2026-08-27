#!/usr/bin/env python3
"""JNTO 訪日外客統計の Excel を読んで JSON にする。

シート = 年（'2026' 等）、行 = 総数/国別、列 = 1 + 月*2（偶数列は前年比の伸率）。
翌月中旬発表で国別まで最新月が揃うため、e-Stat（約3ヶ月遅れ）より優先して使う。

usage: jnto_parse.py <xlsx> <year> <prev_year>
出力: {"months":[..], "cur":[12], "prev":[12], "latestMonth":n, "countriesCur":{}, "countriesPrev":{}}
"""
import json
import sys

import openpyxl

# シート上の表記を、デッキで使う国名に寄せる
NAME = {"英国": "イギリス", "米国": "アメリカ", "豪州": "オーストラリア"}
MAJOR = ["韓国", "中国", "台湾", "香港", "タイ", "シンガポール", "マレーシア", "フィリピン",
         "インドネシア", "ベトナム", "インド", "オーストラリア", "アメリカ", "カナダ",
         "メキシコ", "イギリス", "フランス", "ドイツ"]


def sheet_map(wb, year):
    ws = wb[str(year)]
    rows = {}
    for r in range(5, ws.max_row + 1):
        lab = ws.cell(r, 1).value or ws.cell(r, 2).value
        if not lab:
            continue
        lab = str(lab).strip()
        lab = NAME.get(lab, lab)
        if lab in rows:          # 先に出てくる行（計）を優先し、内訳の重複名で上書きしない
            continue
        rows[lab] = {m: ws.cell(r, 1 + m * 2).value for m in range(1, 13)}
    return rows


def main():
    xlsx, year, prev_year = sys.argv[1], int(sys.argv[2]), int(sys.argv[3])
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    cur_s, prv_s = sheet_map(wb, year), sheet_map(wb, prev_year)

    cur = [cur_s["総数"][m] for m in range(1, 13)]
    prev = [prv_s["総数"][m] for m in range(1, 13)]
    latest = max((m for m in range(1, 13) if cur[m - 1] is not None), default=0)
    if latest == 0:
        raise SystemExit("JNTO: 当年の月次総数が1件も取れませんでした（シート構造の変更を疑う）")

    missing = [c for c in MAJOR if cur_s.get(c, {}).get(latest) is None or prv_s.get(c, {}).get(latest) is None]
    out = {
        "months": [f"{m}月" for m in range(1, 13)],
        "cur": cur,
        "prev": prev,
        "latestMonth": latest,
        "countriesCur": {c: cur_s[c][latest] for c in MAJOR if c not in missing},
        "countriesPrev": {c: prv_s[c][latest] for c in MAJOR if c not in missing},
        "missingCountries": missing,
    }
    json.dump(out, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
